from io import BytesIO
from typing import Any, Dict, Tuple

from openpyxl import Workbook

from src.models.constants import PROFISSIONAIS
from src.services.utils_service import UtilsService


class ExcelService:
    """
    Serviço para geração de arquivos Excel estruturados.
    """

    def __init__(self, dados: Dict[str, Any], resp_selecionado: str):
        self.dados = dados
        self.resp_selecionado = resp_selecionado
        self.profissional = PROFISSIONAIS[self.resp_selecionado]
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "DADOS_IA"

    def _build_base_map(self) -> list:
        d = self.dados
        p = self.profissional
        return [
            ("proponente", str(d.get("proponente", "")).upper()),
            ("cpf_cnpj", str(d.get("cpf_cnpj", ""))),
            ("ddd", str(d.get("ddd", ""))),
            ("telefone", str(d.get("telefone", ""))),
            ("endereco_literal", str(d.get("endereco_literal", "")).upper()),
            ("coordenada_s", str(d.get("coordenada_s", ""))),
            ("coordenada_w", str(d.get("coordenada_w", ""))),
            ("complemento", str(d.get("complemento", "")).upper()),
            ("bairro", str(d.get("bairro", "")).upper()),
            ("cep", str(d.get("cep", ""))),
            ("municipio", str(d.get("municipio", "")).upper()),
            ("uf", str(d.get("uf", ""))),
            ("valor_terreno", UtilsService.to_f(d.get("valor_terreno", 0))),
            ("matricula", str(d.get("matricula", ""))),
            ("oficio", str(d.get("oficio", ""))),
            ("comarca", str(d.get("comarca", ""))),
            ("uf_matricula", str(d.get("uf_matricula", ""))),
            ("categoria", "Casa"),
            ("uso", "Residencial"),
            ("finalidade_vistoria", "Vistoria para aferição de obra"),
            ("valor_imovel", UtilsService.to_f(d.get("valor_imovel", 0))),
            ("numero_etapas", UtilsService.to_f(d.get("numero_etapas", 0))),
            ("empresa", p["empresa"].upper()),
            ("cnpj", p["cnpj"]),
            ("cpf_empresa", p["cpf_emp"]),
            ("nome_responsavel", p["nome_resp"].upper()),
            ("cpf_responsavel", p["cpf_resp"]),
            ("registro", p["registro"].upper()),
        ]

    def _fill_basic_data(self) -> None:
        mapa = self._build_base_map()
        for row_idx, (rotulo, valor) in enumerate(mapa, start=1):
            self.ws.cell(row=row_idx, column=1, value=rotulo)
            self.ws.cell(row=row_idx, column=2, value=valor)

    def _fill_incidencias(self) -> None:
        incs = self.dados.get("incidencias", [])
        for i in range(20):
            row = i + 1
            self.ws.cell(row=row, column=3, value=f"incidencia_{i + 1}")
            self.ws.merge_cells(
                start_row=row, start_column=4, end_row=row, end_column=6
            )
            self.ws.cell(
                row=row,
                column=4,
                value=UtilsService.to_f(incs[i]) if i < len(incs) else 0.0,
            )

    def _fill_acumulado_proposto(self) -> None:
        ap = self.dados.get("acumulado_proposto", [])
        stop = False
        for i in range(37):
            row = i + 1
            self.ws.cell(row=row, column=7, value=f"acumulado_{i}")
            self.ws.merge_cells(
                start_row=row, start_column=8, end_row=row, end_column=10
            )
            if not stop and i < len(ap):
                v = UtilsService.to_f(ap[i])
                self.ws.cell(row=row, column=8, value=v)
                if v >= 100:
                    stop = True
            else:
                self.ws.cell(row=row, column=8, value=None)

    def generate(self) -> Tuple[bytes, str]:
        """
        Gera um arquivo Excel DADOS_IA a partir dos dados extraídos pelo Gemini.

        Retorna (excel_bytes, nome_base_proponente).
        """
        self._fill_basic_data()
        self._fill_incidencias()
        self._fill_acumulado_proposto()

        out = BytesIO()
        self.wb.save(out)

        proponente_str = str(self.dados.get("proponente", "")).strip()
        primeiro_nome = proponente_str.split()[0].upper() if proponente_str else "LAUDO"
        return out.getvalue(), primeiro_nome
