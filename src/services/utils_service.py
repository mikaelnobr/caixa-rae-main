import os
import re
from datetime import datetime
from typing import Any, Optional

from openpyxl.cell.cell import MergedCell


class UtilsService:
    """
    Serviço com métodos utilitários gerais.
    """

    @staticmethod
    def get_secret(key: str) -> Optional[str]:
        """
        Busca chaves no secrets ou variáveis de ambiente, limpando aspas e espaços.
        """
        try:
            import streamlit as st

            if key in st.secrets:
                return st.secrets[key]
        except Exception:
            pass
        val = os.getenv(key) or os.getenv(key.upper())
        if val:
            return val.strip().strip("'").strip('"')
        return None

    @staticmethod
    def to_f(v: Any) -> float:
        """
        Converte qualquer valor para float de forma segura.
        """
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        clean = re.sub(r"[^\d,.-]", "", str(v))
        if not clean:
            return 0.0
        try:
            return float(clean.replace(",", "."))
        except Exception:
            return 0.0

    @staticmethod
    def calcular_serial_data(data_str: Optional[str]) -> int:
        """
        Calcula o serial da data conforme regra: (Ano - 2024) * 12 + Mês + 288.
        """
        try:
            if not data_str or "/" not in data_str:
                dt = datetime.now()
            else:
                dt = datetime.strptime(data_str.strip(), "%d/%m/%Y")

            return (dt.year - 2024) * 12 + dt.month + 288
        except Exception:
            dt = datetime.now()
            return (dt.year - 2024) * 12 + dt.month + 288

    @staticmethod
    def safe_write(ws: Any, coord: str, val: Any) -> None:
        """
        Escreve um valor em uma célula do worksheet, tratando células mescladas.
        """
        try:
            if val is None:
                return

            cell = ws[coord]

            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if coord in merged_range:
                        ws.cell(
                            row=merged_range.min_row, column=merged_range.min_col
                        ).value = val
                        return

            ws[coord].value = val
        except Exception:
            pass

    @staticmethod
    def dmstodecimal(coord_str: Any) -> str:
        """
        Converte formato DMS (Graus, Minutos, Segundos) para Decimal.
        Ex: 23º32'51,234" -> "-23.547565"
        """
        if not coord_str:
            return "0.0"

        s_coord = str(coord_str).strip()

        # Tenta converter direto se já for decimal
        try:
            if re.fullmatch(r"^-?\d+([.,]\d+)?$", s_coord):
                return s_coord.replace(",", ".")
        except Exception:
            pass

        # Extrai apenas os números (incluindo decimais)
        # Ignora símbolos como º, ', ", etc.
        parts = re.findall(r"(\d+(?:[.,]\d+)?)", s_coord)

        if not parts:
            return "0.0"

        try:
            deg = float(parts[0].replace(",", "."))
            mins = float(parts[1].replace(",", ".")) if len(parts) > 1 else 0.0
            sec = float(parts[2].replace(",", ".")) if len(parts) > 2 else 0.0

            decimal = deg + (mins / 60.0) + (sec / 3600.0)

            # Assumimos negativo por padrão, a menos que indique N, E ou L (Leste).
            is_positive = any(c in s_coord.upper() for c in ["N", "E", "L"])
            if not is_positive:
                decimal = -decimal

            # Retorna como string garantindo o ponto
            return str(round(decimal, 8)).replace(",", ".")
        except Exception:
            return "0.0"

    @staticmethod
    def get_now_timestamp() -> str:
        """
        Retorna o timestamp atual no formato DDMMYYYYHHMMSS
        """
        return datetime.now().strftime("%d%m%Y%H%M%S")
