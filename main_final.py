import streamlit as st
import sys
import os
import re
import json
import time
import tempfile
import zipfile
import gc  # Gestão de memória
from datetime import datetime
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# --- CARREGAMENTO DO AMBIENTE ---
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

def get_secret(key):
    """Busca chaves no secrets ou variáveis de ambiente, limpando aspas e espaços."""
    try:
        if key in st.secrets: return st.secrets[key]
    except: pass
    val = os.getenv(key) or os.getenv(key.upper())
    if val:
        return val.strip().strip("'").strip('"')
    return None

def calcular_serial_data(data_str):
    """Calcula o serial da data conforme regra: (Ano - 2024) * 12 + Mês + 288."""
    try:
        if not data_str or "/" not in data_str:
            dt = datetime.now()
        else:
            # Tenta converter data do laudo (DD/MM/AAAA)
            dt = datetime.strptime(data_str.strip(), "%d/%m/%Y")
        
        return (dt.year - 2024) * 12 + dt.month + 288
    except:
        dt = datetime.now()
        return (dt.year - 2024) * 12 + dt.month + 288

# --- CONFIGURAÇÃO INICIAL ---
st.set_page_config(page_title="Automação RAE CAIXA", page_icon="🏛️", layout="centered")

# --- BASE DE DADOS DE PROFISSIONAIS ---
PROFISSIONAIS = {
    "FRANCISCO DAVID MENESES DOS SANTOS": {
        "empresa": "FRANCISCO DAVID MENESES DOS SANTOS",
        "cnpj": "54.801.096/0001-16",
        "cpf_emp": "058.756.003-73",
        "nome_resp": "FRANCISCO DAVID MENESES DOS SANTOS",
        "cpf_resp": "058.756.003-73",
        "registro": "336241CE"
    },
    "PALLOMA TEIXEIRA DA SILVA": {
        "empresa": "PALLOMA TEIXEIRA DA SILVA",
        "cnpj": "54.862.474/0001-71",
        "cpf_emp": "064.943.593-10",
        "nome_resp": "PALLOMA TEIXEIRA DA SILVA",
        "cpf_resp": "064.943.593-10",
        "registro": "A184355-9"
    },
    "SANDY PEREIRA CORDEIRO": {
        "empresa": "SANDY PEREIRA CORDEIRO - CS ENGENHARIA",
        "cnpj": "54.794.898/0001-46",
        "cpf_emp": "071.222.553-60",
        "nome_resp": "SANDY PEREIRA CORDEIRO",
        "cpf_resp": "071.222.553-60",
        "registro": "356882CE"
    },
    "TIAGO VICTOR DE SOUSA": {
        "empresa": "TIAGO VICTOR DE SOUSA - T V S ENGENHARIA E ASSESSORIA",
        "cnpj": "54.806.521/0001-60",
        "cpf_emp": "068.594.803-00",
        "nome_resp": "TIAGO VICTOR DE SOUSA",
        "cpf_resp": "068.594.803-00",
        "registro": "346856CE"
    },
    "DAVID ARRUDA VIANA": {
        "empresa": "DAVID ARRUDA VIANA - DAV ENGENHARIA LTDA",
        "cnpj": "51.508.674/0001-32",
        "cpf_emp": "033.467.853-60",
        "nome_resp": "DAVID ARRUDA VIANA",
        "cpf_resp": "033.467.853-60",
        "registro": "061931352-8"
    },
    "MIKAELL GUSTAVO FARIAS GOMES": {
        "empresa": "MIKAELL GUSTAVO FARIAS GOMES",
        "cnpj": "51.899.957/0001-52",
        "cpf_emp": "057.953.393-00",
        "nome_resp": "MIKAELL GUSTAVO FARIAS GOMES",
        "cpf_resp": "057.953.393-00",
        "registro": "367577CE"
    },
    "YAN LUCAS E SILVA VASCONCELOS": {
        "empresa": "YAN LUCAS E SILVA VASCONCELOS",
        "cnpj": "54.732.603/0001-07",
        "cpf_emp": "038.621.633-93",
        "nome_resp": "YAN LUCAS E SILVA VASCONCELOS",
        "cpf_resp": "038.621.633-93",
        "registro": "365506"
    }
}
# --- DEPENDÊNCIAS ---
try:
    import pdfplumber
    import google.generativeai as genai
    import gspread
    from google.oauth2.service_account import Credentials
    DEPENDENCIAS_OK = True
except ImportError as e:
    DEPENDENCIAS_OK = False
    ERRO_IMPORT = str(e)

# --- UTILITÁRIOS ---
def to_f(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    clean = re.sub(r'[^\d,.-]', '', str(v))
    if not clean: return 0.0
    try: return float(clean.replace(',', '.'))
    except: return 0.0

def safe_write(ws, coord, val):
    try:
        # 🚫 NUNCA escrever None
        if val is None:
            return

        cell = ws[coord]

        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if coord in merged_range:
                    ws.cell(
                        row=merged_range.min_row,
                        column=merged_range.min_col
                    ).value = val
                    return

        ws[coord].value = val
    except:
        pass

# --- AUTENTICAÇÃO GOOGLE SILENCIOSA ---
def get_gspread_client():
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    if os.path.exists("service_account.json"):
        try: return gspread.authorize(Credentials.from_service_account_file("service_account.json", scopes=scopes))
        except: pass
    json_str = get_secret("GCP_SERVICE_ACCOUNT")
    if json_str:
        try:
            clean_str = json_str.strip().replace('\\\\n', '\n').replace('\\n', '\n')
            info = json.loads(clean_str, strict=False)
            if "private_key" in info: info["private_key"] = info["private_key"].replace('\\n', '\n')
            return gspread.authorize(Credentials.from_service_account_info(info, scopes=scopes))
        except: pass
    return None

# --- SINCRONIZAÇÃO (43 COLUNAS ALINHADAS COM O CSV) ---
def save_to_google_sheets(dados_ia, resp_nome):
    try:
        client = get_gspread_client()
        if not client: return False, "Falha na autenticação do Google Sheets (service_account.json faltando ou inválido)."
        
        sheet_id = get_secret("GOOGLE_SHEET_ID")
        if not sheet_id: return False, "A configuração GOOGLE_SHEET_ID não foi encontrada (adicione no .env ou st.secrets)."
        
        sheet = client.open_by_key(sheet_id).get_worksheet(0)
        tel = f"({dados_ia.get('ddd','')}) {dados_ia.get('telefone','')}"
        val_data_evento = calcular_serial_data(dados_ia.get("data_referencia", ""))
        
        v_terreno = to_f(dados_ia.get("valor_terreno", 0))
        a_terreno = to_f(dados_ia.get("area_terreno", 0))
        v_unit_terreno = round(v_terreno / a_terreno, 2) if a_terreno > 0 else 0
        
        endereco_literal = str(dados_ia.get("endereco_literal", "")).upper()
        
        # Mapeamento rigoroso conforme "dados caixa - Página1.csv"
        row = [
            str(dados_ia.get("empresa_responsavel", "")).upper(), # 1: Responsável pela avaliação
            str(resp_nome).upper(),                             # 2: Representante legal
            str(dados_ia.get("proponente", "")).upper(),         # 3: Informante
            tel,                                                 # 4: Telefone
            endereco_literal,                                    # 5: Endereço
            str(dados_ia.get("bairro", "")).upper(),             # 6: Bairro
            str(dados_ia.get("municipio", "")).upper(),          # 7: Município
            str(dados_ia.get("coordenada_s", "")),               # 8: Latitude (GMS)
            str(dados_ia.get("coordenada_w", "")),               # 9: Longitude (GMS)
            a_terreno,                                           # 10: Área Terreno
            to_f(dados_ia.get("area_construida", 0)),            # 11: Área Construída
            to_f(dados_ia.get("quartos", 0)),                    # 12: Quartos
            to_f(dados_ia.get("banheiros", 0)),                  # 13: Banheiros
            to_f(dados_ia.get("suites", 0)),                     # 14: Banheiro Privativo (Suites)
            to_f(dados_ia.get("vagas", 0)),                      # 15: Vagas
            to_f(dados_ia.get("valor_imovel", 0)),               # 16: Valor Global
            to_f(dados_ia.get("valor_unitario", 0)),             # 17: Valor Unitário
            str(dados_ia.get("padrao_acabamento", "")).upper(),  # 18: Padrão Acabamento
            str(dados_ia.get("estado_conservacao", "")).upper(), # 19: Estado Conservação
            str(dados_ia.get("infraestrutura", "")).upper(),     # 20: Infraestrutura
            str(dados_ia.get("servicos_publicos", "")).upper(),   # 21: Serviços Públicos
            str(dados_ia.get("usos_predominantes", "")).upper(), # 22: Usos Predominantes
            str(dados_ia.get("via_acesso", "")).upper(),         # 23: Via de Acesso
            str(dados_ia.get("regiao_contexto", "")).upper(),    # 24: Região Contexto
            str(dados_ia.get("idade_estimada", "")),             # 25: Idade Estimada
            val_data_evento,                                     # 26: Data Evento
            # --- BLOCO TERRENO (Repetidos conforme planilha) ---
            str(dados_ia.get("proponente", "")).upper(),         # 27: Informante (T)
            tel,                                                 # 28: Telefone (T)
            endereco_literal,                                    # 29: Endereço (T)
            str(dados_ia.get("municipio", "")).upper(),          # 30: Município (T)
            str(dados_ia.get("bairro", "")).upper(),             # 31: Bairro (T)
            str(dados_ia.get("coordenada_s", "")),               # 32: Latitude (T)
            str(dados_ia.get("coordenada_w", "")),               # 33: Longitude (T)
            a_terreno,                                           # 34: Área (T)
            to_f(dados_ia.get("testada", 0)),                    # 35: Testada
            v_terreno,                                           # 36: Valor
            v_unit_terreno,                                      # 37: Valor unitário
            str(dados_ia.get("infraestrutura", "")).upper(),     # 38: Infraestrutura (T)
            str(dados_ia.get("servicos_publicos", "")).upper(),   # 39: Serviços Públicos (T)
            str(dados_ia.get("usos_predominantes", "")).upper(), # 40: Usos Predominantes (T)
            str(dados_ia.get("via_acesso", "")).upper(),         # 41: Via de Acesso (T)
            str(dados_ia.get("regiao_contexto", "")).upper(),    # 42: Região Contexto (T)
            val_data_evento                                      # 43: Data Evento (T)
        ]
        sheet.append_row(row)
        return True, "Salvo com sucesso no Google Sheets!"
    except Exception as e:
        return False, f"Erro ao salvar no Google Sheets: {e}"

# --- MOTOR DE EXTRAÇÃO VISUAL ---
def get_text_with_layout(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # layout=True mantém o posicionamento visual X,Y em espaços (perfeito para "planilhas" nativas)
            ext = page.extract_text(layout=True)
            if ext: text += ext + "\n"
    return text

def call_gemini(api_key, prompt):
    genai.configure(api_key=api_key)
    schema = {
        "type": "object",
        "properties": {
            "proponente": {"type": "string"}, "cpf_cnpj": {"type": "string"}, "ddd": {"type": "string"},
            "telefone": {"type": "string"}, "endereco_literal": {"type": "string"}, "bairro": {"type": "string"},
            "cep": {"type": "string"}, "municipio": {"type": "string"}, "uf": {"type": "string"},
            "coordenada_s": {"type": "string"}, "coordenada_w": {"type": "string"}, "valor_terreno": {"type": "number"},
            "valor_imovel": {"type": "number"}, "valor_unitario": {"type": "number"}, "testada": {"type": "number"},
            "matricula": {"type": "string"}, "oficio": {"type": "string"}, "comarca": {"type": "string"}, "uf_matricula": {"type": "string"},
            "incidencias": {"type": "array", "items": {"type": "number"}}, "numero_etapas": {"type": "number"},
            "acumulado_proposto": {"type": "array", "items": {"type": "number"}}, "idade_estimada": {"type": "string"},
            "area_terreno": {"type": "number"}, "area_construida": {"type": "number"}, "quartos": {"type": "number"},
            "banheiros": {"type": "number"}, "suites": {"type": "number"}, "vagas": {"type": "number"},
            "padrao_acabamento": {"type": "string"}, "estado_conservacao": {"type": "string"},
            "infraestrutura": {"type": "string"}, "servicos_publicos": {"type": "string"},
            "usos_predominantes": {"type": "string"}, "via_acesso": {"type": "string"}, "regiao_contexto": {"type": "string"},
            "data_referencia": {"type": "string"},
            "empresa_responsavel": {"type": "string"}
        },
        "required": ["proponente", "ddd", "telefone", "endereco_literal", "incidencias", "empresa_responsavel"]
    }
    model = genai.GenerativeModel('gemini-2.5-flash')
    for d in [2, 4, 8]:
        try:
            res = model.generate_content(prompt, generation_config=genai.types.GenerationConfig(response_mime_type="application/json", response_schema=schema, temperature=0.1))
            return json.loads(res.text)
        except Exception as e:
            if "429" in str(e): time.sleep(d); continue
            raise e

def process_single_pdf(pdf_file, api_key, resp_selecionado, status_container):
    """Processa um único PDF e retorna (dados, excel_bytes, nome_arquivo) ou levanta exceção."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(pdf_file.getbuffer()); tp = tmp.name
    try:
        status_container.write(f"📄 Extraindo layout visual de **{pdf_file.name}**...")
        md = get_text_with_layout(tp)
    finally:
        if os.path.exists(tp): os.remove(tp)

    prompt = f"""
    Você é um ENGENHEIRO REVISOR ESPECIALISTA EM LAUDOS DE AVALIAÇÃO DA CAIXA ECONÔMICA FEDERAL.

    Sua tarefa é extrair informações técnicas do laudo abaixo e retornar EXCLUSIVAMENTE um JSON
    válido, rigorosamente conforme o schema fornecido. NÃO explique nada fora do JSON.

    ────────────────────────────────────────
    EXTRAÇÃO DE COORDENADAS GEOGRÁFICAS

    O texto abaixo não é Markdown! Ele é um RETRATO ESPACIAL do PDF. As colunas estão devidamente alinhadas
    com espaços em branco reproduzindo a posição exata da tabela original.

    Olhe o visual do texto perto da palavra "Latitude" e "Longitude Oeste":

    REGRA MATEMÁTICA DEFINITIVA NO BRASIL:
    1. Mapeie visualmente as colunas de [Graus], [Min] e [Seg].
    2. Extraia o primeiro conjunto que encontrar na mesma linha lida horizontalmente.
    3. Extraia o segundo conjunto na mesma linha.
    4. Classifique-os estritamente pelo GRAU:
       - Se o Grau for entre 00º e 33º -> Este grupo inteiro (Grau+Min+Seg) é a LATITUDE (coordenada_s).
       - Se o Grau for entre 34º e 74º -> Este grupo inteiro (Grau+Min+Seg) é a LONGITUDE (coordenada_w).

    FORMATO FINAL EXIGIDO NAS CHAVES JSON:
    coordenada_s: "XXºYY'ZZ,ZZZ\""
    coordenada_w: "XXºYY'ZZ,ZZZ\""
    ────────────────────────────────────────

    REGRAS DE CRONOGRAMA:
    1. Localize a tabela de "Cronograma" ou "Parcelas".
    2. **acumulado_proposto**:
       - O primeiro valor da lista DEVE ser o da linha "Pré-executado".
       - É OBRIGATÓRIO incluir o valor do "Pré-executado" mesmo que seja 0.00.
       - NÃO pule a etapa 0. Se ela for 0%, retorne 0.0 como primeiro item da lista.
       - Em seguida, extraia os valores da coluna "% Acumulado" para as parcelas 1, 2, 3...

    REGRAS DE EXTRAÇÃO DO CRONOGRAMA (INCIDÊNCIAS):

    1. Localize a tabela:
       - "Cronograma Físico-Financeiro"
       - "Discriminação dos Serviços"
       - ou "Orçamento Proposto"

    2. Extraia a coluna de INCIDÊNCIA (ou PESOS).
       - Retorne EXATAMENTE 20 valores percentuais.
       - Preserve a ordem original das etapas.
       - NÃO normalize, NÃO ajuste, NÃO redistribua valores.

    3. Se houver menos de 20 etapas:
       - Complete a lista com 0.0 até atingir 20 itens.

    ────────────────────────────────────────
    REGRAS PARA CAMPOS CLASSIFICÁVEIS:

    - VIA_ACESSO: Retorne SOMENTE se estiver explícito no laudo.
      Valores aceitos: LOCAL, COLETORA ou ARTERIAL.
      Se não estiver explícito, retorne string vazia.

    - PADRAO_ACABAMENTO, ESTADO_CONSERVACAO, REGIAO_CONTEXTO:
      - NÃO infira.
      - NÃO classifique por suposição.
      - Se não estiver textual e claramente descrito, retorne string vazia.

    ────────────────────────────────────────
    OUTRAS REGRAS CRÍTICAS:

    1. ENDERECO_LITERAL:
       - Copie EXATAMENTE como consta na identificação do imóvel.
       - Preserve abreviações, números e ordem.

    2. MATRÍCULA:
       - Extraia número da matrícula.
       - OFÍCIO = número do cartório.
       - COMARCA = município do registro.
       - UF_MATRICULA = estado do cartório.

    3. IDADE_ESTIMADA:
       - Capture o TEXTO LITERAL COMPLETO.
       - Exemplos válidos: "5 anos", "Novo", "Na Planta".

    4. DATA_REFERENCIA:
       - Utilize EXCLUSIVAMENTE a data da AVALIAÇÃO DO IMÓVEL.
       - Ignore datas de ART, vistoria, assinatura ou emissão.
       - Formato obrigatório: DD/MM/AAAA.

    5. EMPRESA_RESPONSAVEL:
       - Vá para a seção final 'SIGNATÁRIOS'.
       - Localize o campo 'Representante legal' associado ao Responsável Técnico.
       - Extraia o nome literal completo do representante legal (Ex: MIKAELL GUSTAVO FARIAS GOMES).

    6. CAMPOS AUSENTES:
       - Se a informação não existir no laudo, retorne:
         - string vazia para textos
         - 0 para números
         - lista vazia quando aplicável

    ERRO GRAVE:
    - Misturar dados entre campos invalida a extração.
    - Inferir informações técnicas não explícitas é proibido.

    ────────────────────────────────────────
    TEXTO DO LAUDO:

    {md}
    """

    status_container.write(f"🤖 Chamando Gemini para **{pdf_file.name}**...")
    dados = call_gemini(api_key, prompt)

    # Gerar arquivo DADOS_IA limpo (NÃO abre o template do usuário)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DADOS_IA"

    p = PROFISSIONAIS[resp_selecionado]

    # === MAPEAMENTO PLANO: Linha -> (Rótulo, Valor) ===
    mapa = [
        # --- DADOS DO PROPONENTE ---
        ("proponente",         str(dados.get("proponente", "")).upper()),
        ("cpf_cnpj",           str(dados.get("cpf_cnpj", ""))),
        ("ddd",                str(dados.get("ddd", ""))),
        ("telefone",           str(dados.get("telefone", ""))),
        # --- ENDEREÇO ---
        ("endereco_literal",   str(dados.get("endereco_literal", "")).upper()),
        ("coordenada_s",       str(dados.get("coordenada_s", ""))),
        ("coordenada_w",       str(dados.get("coordenada_w", ""))),
        ("complemento",        str(dados.get("complemento", "")).upper()),
        ("bairro",             str(dados.get("bairro", "")).upper()),
        ("cep",                str(dados.get("cep", ""))),
        ("municipio",          str(dados.get("municipio", "")).upper()),
        ("uf",                 str(dados.get("uf", ""))),
        # --- VALORES E MATRÍCULA ---
        ("valor_terreno",      to_f(dados.get("valor_terreno", 0))),
        ("matricula",          str(dados.get("matricula", ""))),
        ("oficio",             str(dados.get("oficio", ""))),
        ("comarca",            str(dados.get("comarca", ""))),
        ("uf_matricula",       str(dados.get("uf_matricula", ""))),
        # --- FIXOS ---
        ("categoria",          "Casa"),
        ("uso",                "Residencial"),
        ("finalidade_vistoria","Vistoria para aferição de obra"),
        # --- RAE ---
        ("valor_imovel",       to_f(dados.get("valor_imovel", 0))),
        ("numero_etapas",      to_f(dados.get("numero_etapas", 0))),
        # --- PROFISSIONAL ---
        ("empresa",            p["empresa"].upper()),
        ("cnpj",               p["cnpj"]),
        ("cpf_empresa",        p["cpf_emp"]),
        ("nome_responsavel",   p["nome_resp"].upper()),
        ("cpf_responsavel",    p["cpf_resp"]),
        ("registro",           p["registro"].upper()),
    ]

    # Escrever rótulos e valores (A1:B28)
    for row_idx, (rotulo, valor) in enumerate(mapa, start=1):
        ws.cell(row=row_idx, column=1, value=rotulo)
        ws.cell(row=row_idx, column=2, value=valor)

    # --- INCIDÊNCIAS (20 valores) na coluna D (D1:D20) com rótulo na coluna C ---
    incs = dados.get("incidencias", [])
    for i in range(20):
        ws.cell(row=i+1, column=3, value=f"incidencia_{i+1}")
        ws.cell(row=i+1, column=4, value=to_f(incs[i]) if i < len(incs) else 0.0)

    # --- ACUMULADO PROPOSTO (até 37 valores) na coluna F (F1:F37) com rótulo na coluna E ---
    ap = dados.get("acumulado_proposto", [])
    stop = False
    for i in range(37):
        ws.cell(row=i+1, column=5, value=f"acumulado_{i}")
        if not stop and i < len(ap):
            v = to_f(ap[i])
            ws.cell(row=i+1, column=6, value=v)
            if v >= 100: stop = True
        else:
            ws.cell(row=i+1, column=6, value=None)

    out = BytesIO()
    wb.save(out)
    processed_data = out.getvalue()

    nome_base = os.path.splitext(pdf_file.name)[0]
    return dados, processed_data, nome_base


def main():
    st.title("🏛️ Automação RAE CAIXA")
    if "processed" not in st.session_state: st.session_state["processed"] = False
    if not DEPENDENCIAS_OK: st.error(f"Erro: {ERRO_IMPORT}"); return

    with st.sidebar:
        st.header("⚙️ Configurações")
        api_key = st.text_input("Gemini API Key:", type="password")
        resp_selecionado = st.selectbox("Responsável Técnico:", options=list(PROFISSIONAIS.keys()))
        if st.session_state["processed"]:
            if st.button("🔄 NOVA FILA"): st.session_state["processed"] = False; st.rerun()

    if st.session_state["processed"]:
        st.success("✅ Processamento Concluído!")
        # Mostrar resultados salvos na sessão
        if "batch_results" in st.session_state:
            res = st.session_state["batch_results"]
            st.info(f"📊 **{res['ok']}** laudos processados com sucesso, **{res['fail']}** falharam de **{res['total']}** total.")
            if res.get("zip_bytes"):
                st.download_button("📥 BAIXAR TODOS OS RAEs (.zip)", res["zip_bytes"], "RAEs_processados.zip", "application/zip")
            if res.get("erros"):
                with st.expander("⚠️ Laudos com Erro"):
                    for err in res["erros"]:
                        st.warning(err)
        return

    pdf_files = st.file_uploader("📄 Laudos Técnicos (PDFs)", type=["pdf"], accept_multiple_files=True)
    gerar_excel = st.checkbox("📊 Gerar arquivo DADOS_IA para RAE", value=True)

    if pdf_files:
        st.caption(f"📁 {len(pdf_files)} laudo(s) selecionado(s)")

    if st.button("🚀 INICIAR PROCESSAMENTO"):
        if not api_key or not pdf_files: st.warning("Preencha a API Key e selecione ao menos 1 PDF."); return

        total = len(pdf_files)
        progress_bar = st.progress(0, text=f"Iniciando processamento de {total} laudo(s)...")
        resultados_excel = []  # lista de (nome_arquivo, bytes)
        erros = []
        ok_count = 0

        for idx, pdf in enumerate(pdf_files):
            progress_bar.progress((idx) / total, text=f"📄 Processando {idx+1}/{total}: {pdf.name}")
            status_container = st.status(f"[{idx+1}/{total}] {pdf.name}", expanded=(idx == len(pdf_files) - 1))
            try:
                dados, excel_bytes, nome_base = process_single_pdf(
                    pdf, api_key, resp_selecionado, status_container
                )

                # Sincronização Google Sheets
                sheets_ok, sheets_msg = save_to_google_sheets(dados, PROFISSIONAIS[resp_selecionado]["nome_resp"])
                if not sheets_ok:
                    status_container.write(f"⚠️ Google Sheets: {sheets_msg}")
                else:
                    status_container.write(f"✅ {sheets_msg}")

                if gerar_excel:
                    primeiro_nome = str(dados.get("proponente", "LAUDO")).split()[0].upper()
                    resultados_excel.append((f"RAE_{primeiro_nome}.xlsx", excel_bytes))

                status_container.update(label=f"✅ [{idx+1}/{total}] {pdf.name}", state="complete")
                ok_count += 1

            except Exception as e:
                erros.append(f"❌ {pdf.name}: {e}")
                status_container.update(label=f"❌ [{idx+1}/{total}] {pdf.name}", state="error")

            gc.collect()

        progress_bar.progress(1.0, text=f"✅ Concluído! {ok_count}/{total} laudos processados.")

        # Empacotar todos os Excel em ZIP
        zip_bytes = None
        if resultados_excel:
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for nome, dados_xl in resultados_excel:
                    zf.writestr(nome, dados_xl)
            zip_bytes = zip_buffer.getvalue()

        # Salvar resultados na sessão para persistir após rerun
        st.session_state["batch_results"] = {
            "ok": ok_count,
            "fail": len(erros),
            "total": total,
            "zip_bytes": zip_bytes,
            "erros": erros
        }
        st.session_state["processed"] = True
        gc.collect()
        st.rerun()


if __name__ == "__main__": main()
